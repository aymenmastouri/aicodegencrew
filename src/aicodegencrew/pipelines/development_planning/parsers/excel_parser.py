"""
Excel Parser for XLSX/XLS files.
"""

from pathlib import Path
from typing import Dict, Any, List


def parse_excel(file_path: Path) -> Dict[str, Any]:
    """
    Parse Excel file and extract data.

    Args:
        file_path: Path to Excel file

    Returns:
        Dictionary with sheets and data
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel parsing. "
            "Install with: pip install openpyxl"
        )

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    result = {
        'sheets': {},
        'metadata': {
            'sheet_names': workbook.sheetnames,
            'sheet_count': len(workbook.sheetnames),
        }
    }

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Extract data
        data = []
        for row in sheet.iter_rows(values_only=True):
            # Skip empty rows
            if any(cell is not None for cell in row):
                data.append(list(row))

        result['sheets'][sheet_name] = {
            'name': sheet_name,
            'data': data,
            'row_count': len(data),
        }

    return result
